from __future__ import annotations

import argparse
import importlib.util
import json
import re
from collections import Counter
from pathlib import Path

import pandas as pd


INVENTORY_FILE = "库存分析看板源数据_含宏.xlsm"
DEVICE_FILE = "2026年设备发货明细-共享版.xlsm"
DEVICE_FILE_ALIASES = ("2026年设备发货明细-共享版.xlsm", "2026年设备发货明细 -共享版.xlsm")
FREIGHT_FILE = "物料运费分析.xlsx"
AFTER_SALES_FILE = "售后跟进.xlsx"
DEVELOPMENT_FILE = "物料开发进度跟进表.xlsx"

DEVICE_REQUIRED_COLUMNS = {
    "设备名称清洗", "销售部门清洗-大类", "总货值", "数量", "发货状态",
}


def js_assignment(name: str, value: object, spaced: bool = False) -> str:
    separator = " = " if spaced else "="
    return f"window.{name}{separator}{json.dumps(value, ensure_ascii=False, separators=(',', ':'))};\n"


def load_legacy_updater(source_dir: Path):
    path = source_dir / "update_dashboard_data.py"
    spec = importlib.util.spec_from_file_location("material_dashboard_legacy_updater", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载更新程序：{path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def clean_text(value: object, fallback: str = "") -> str:
    if pd.isna(value):
        return fallback
    text = str(value).strip()
    return text or fallback


def number_series(frame: pd.DataFrame, column: str) -> pd.Series:
    return pd.to_numeric(frame[column], errors="coerce").fillna(0.0)


def normalize_freight_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """兼容物料运费分析表的新旧字段名称。"""
    frame = frame.copy()
    aliases = {
        "货值": ("货值", "货值（供应链单价*数量）", "货值(供应链单价*数量)"),
        "供应链单价": ("供应链单价", "供应链单价(推测)", "供应链单价（推测）"),
    }
    for canonical, candidates in aliases.items():
        if canonical in frame.columns:
            continue
        matched = next((column for column in candidates if column in frame.columns), None)
        if matched:
            frame[canonical] = frame[matched]
    if "货值" not in frame.columns and {"供应链单价", "总数量"}.issubset(frame.columns):
        frame["货值"] = (
            pd.to_numeric(frame["供应链单价"], errors="coerce").fillna(0.0)
            * pd.to_numeric(frame["总数量"], errors="coerce").fillna(0.0)
        )
    return frame


def normalize_province(value: object) -> str:
    text = clean_text(value)
    for suffix in (
        "壮族自治区", "回族自治区", "维吾尔自治区", "特别行政区", "自治区", "省", "市"
    ):
        if text.endswith(suffix):
            return text[: -len(suffix)].strip()
    return text


def resolve_device_path(source_dir: Path) -> Path:
    for filename in DEVICE_FILE_ALIASES:
        candidate = source_dir / filename
        if candidate.exists():
            return candidate
    return source_dir / DEVICE_FILE


def load_device_workbook(device_path: Path) -> tuple[pd.DataFrame, list[dict[str, object]]]:
    workbook = pd.ExcelFile(device_path)
    frames = []
    sheets = []
    for raw_sheet_name in workbook.sheet_names:
        frame = pd.read_excel(device_path, sheet_name=raw_sheet_name)
        missing = DEVICE_REQUIRED_COLUMNS.difference(frame.columns)
        if missing:
            continue
        sheet_name = raw_sheet_name.strip()
        frame = frame.copy()
        category = frame["销售部门清洗-大类"].map(clean_text)
        province_column = next(
            (column for column in ("省份清洗", "省区清洗") if column in frame.columns),
            None,
        )
        province = (
            frame[province_column].map(normalize_province)
            if province_column
            else pd.Series("", index=frame.index, dtype="object")
        )
        is_mingmang = sheet_name.startswith("鸣忙-") | category.eq("鸣忙")
        if province_column is None and not bool(is_mingmang.all()):
            continue
        frame["销售部门清洗-省区"] = province
        frame.loc[is_mingmang, "销售部门清洗-省区"] = "鸣忙"
        fallback_region = category.where(~category.isin(["", "省区", "鸣忙"]), "未分区")
        blank_region = frame["销售部门清洗-省区"].eq("")
        frame.loc[blank_region, "销售部门清洗-省区"] = fallback_region.loc[blank_region]
        frame["实际发货状态"] = frame["发货状态"].map(clean_text)
        frame["实际发货数量"] = number_series(frame, "数量")
        frame["数量"] = number_series(frame, "数量")
        frame["总货值"] = number_series(frame, "总货值")
        frame["设备名称清洗"] = frame["设备名称清洗"].map(lambda value: clean_text(value, "未命名设备"))
        order_date_column = next(
            (column for column in ("下单日期", "下单时间") if column in frame.columns),
            None,
        )
        if order_date_column is None:
            continue
        frame["下单时间标准"] = pd.to_datetime(frame[order_date_column], errors="coerce")
        valid_order_date = frame["下单时间标准"].notna()
        frame.loc[valid_order_date, "月份"] = frame.loc[valid_order_date, "下单时间标准"].dt.month
        product_columns = [column for column in ("商品名称", "设备名称", "冰柜需求型号") if column in frame.columns]
        if product_columns:
            product = frame[product_columns[0]].map(clean_text)
            for column in product_columns[1:]:
                product = product.where(product.ne(""), frame[column].map(clean_text))
            frame["产品"] = product.where(product.ne(""), frame["设备名称清洗"])
        else:
            frame["产品"] = frame["设备名称清洗"]
        frame["来源Sheet"] = sheet_name
        frames.append(frame)
        sheets.append({
            "sheet": sheet_name,
            "rows": len(frame),
            "shippedRows": int(frame["实际发货状态"].eq("已发货").sum()),
        })
    if not frames:
        raise ValueError(f"{device_path.name} 中没有找到包含设备字段的工作表")
    return pd.concat(frames, ignore_index=True, sort=False), sheets


def build_device_outbound_file(source_dir: Path, device_path: Path) -> dict[str, object]:
    device_df, sheets = load_device_workbook(device_path)
    device_df = device_df.loc[
        device_df["下单时间标准"].notna()
        & device_df["销售部门清洗-省区"].ne("")
        & device_df["实际发货状态"].eq("已发货")
    ].copy()
    device_df["下单日期"] = device_df["下单时间标准"].dt.strftime("%Y-%m-%d")
    summary = {}
    for order_date, group in device_df.groupby("下单日期"):
        regions = (
            group.groupby("销售部门清洗-省区")
            .agg(amount=("总货值", "sum"), quantity=("实际发货数量", "sum"))
            .reset_index()
            .sort_values("amount", ascending=False)
        )
        summary[order_date] = []
        for _, region in regions.iterrows():
            province = str(region["销售部门清洗-省区"])
            province_group = group.loc[group["销售部门清洗-省区"].eq(province)]
            devices = (
                province_group.groupby("设备名称清洗")
                .agg(amount=("总货值", "sum"), quantity=("实际发货数量", "sum"))
                .reset_index()
                .sort_values("amount", ascending=False)
            )
            summary[order_date].append({
                "province": province,
                "amount": round(float(region["amount"]), 2),
                "quantity": round(float(region["quantity"]), 2),
                "devices": [
                    {
                        "name": str(row["设备名称清洗"]),
                        "amount": round(float(row["amount"]), 2),
                        "quantity": round(float(row["quantity"]), 2),
                    }
                    for _, row in devices.iterrows()
                ],
            })
    (source_dir / "device_outbound_data.js").write_text(
        js_assignment("DEVICE_OUTBOUND_DATA", summary, spaced=True), encoding="utf-8"
    )
    return {
        "sheets": sheets,
        "dateRange": [min(summary), max(summary)] if summary else [],
        "dates": len(summary),
        "shippedRows": len(device_df),
        "quantity": round(float(device_df["实际发货数量"].sum()), 2),
        "amount": round(float(device_df["总货值"].sum()), 2),
    }


def build_freight_detail_files(source_dir: Path, freight_path: Path) -> dict[str, int]:
    frame = normalize_freight_columns(
        pd.read_excel(freight_path, sheet_name="物料运费分析")
    )
    required = {
        "出库单号", "主运单号", "物料名称", "总数量", "运费", "货值", "月份",
        "发货仓库", "市", "供应链单价", "物料起发量",
    }
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError("物料运费分析缺少明细/试算字段：" + "、".join(sorted(missing)))

    frame = frame.copy()
    frame["月份"] = pd.to_numeric(frame["月份"], errors="coerce")
    frame = frame.loc[frame["月份"].notna()].copy()
    frame["月份"] = frame["月份"].astype(int)
    for column in ("总数量", "运费", "货值", "供应链单价", "物料起发量"):
        frame[column] = number_series(frame, column)
    for column in ("出库单号", "主运单号", "物料名称", "发货仓库", "市"):
        frame[column] = frame[column].map(clean_text)
    frame.loc[frame["物料名称"].eq(""), "物料名称"] = "未命名物料"

    order_details = []
    for _, group in frame.groupby(["月份", "出库单号"], sort=False, dropna=False):
        goods_value = float(group["货值"].sum())
        freight = float(group["运费"].max())
        materials = "、".join(dict.fromkeys(group["物料名称"].tolist()))
        waybill = next((value for value in group["主运单号"] if value), clean_text(group.iloc[0]["出库单号"]))
        order_details.append([
            int(group.iloc[0]["月份"]),
            waybill,
            materials,
            round(float(group["总数量"].sum()), 2),
            round(freight, 2),
            round(freight / goods_value * 100 if goods_value else 0.0, 4),
        ])
    (source_dir / "material_freight_order_details.js").write_text(
        js_assignment("FREIGHT_ORDER_DETAILS", order_details, spaced=True), encoding="utf-8"
    )

    materials = sorted(frame["物料名称"].drop_duplicates().tolist())
    material_index = {name: index for index, name in enumerate(materials)}
    material_rows = [
        [
            int(row["月份"]), clean_text(row["出库单号"]), material_index[row["物料名称"]],
            round(float(row["总数量"]), 2), round(float(row["运费"]), 2),
            round(float(row["货值"]), 2),
            round(float(row["运费"]) / float(row["货值"]) * 100 if float(row["货值"]) else 0.0, 6),
        ]
        for _, row in frame.iterrows()
    ]
    (source_dir / "material_freight_material_details.js").write_text(
        js_assignment("FREIGHT_MATERIAL_SOURCE", {"materials": materials, "rows": material_rows}),
        encoding="utf-8",
    )

    cities = sorted(frame["市"].drop_duplicates().tolist())
    warehouses = sorted(frame["发货仓库"].drop_duplicates().tolist())
    orders = sorted(frame["出库单号"].drop_duplicates().tolist())
    city_index = {name: index for index, name in enumerate(cities)}
    warehouse_index = {name: index for index, name in enumerate(warehouses)}
    order_index = {name: index for index, name in enumerate(orders)}
    material_meta = []
    for material in materials:
        group = frame.loc[frame["物料名称"].eq(material)]
        prices = group.loc[group["供应链单价"].gt(0), "供应链单价"]
        starts = group.loc[group["物料起发量"].gt(0), "物料起发量"]
        material_meta.append([
            round(float(prices.mode().iloc[0] if not prices.empty else 0.0), 4),
            round(float(starts.mode().iloc[0] if not starts.empty else 0.0), 2),
        ])
    calculator_rows = [
        [
            int(row["月份"]), material_index[row["物料名称"]], city_index[row["市"]],
            warehouse_index[row["发货仓库"]], order_index[row["出库单号"]],
            round(float(row["总数量"]), 2), round(float(row["运费"]), 2),
            round(float(row["运费"]) / float(row["货值"]) * 100 if float(row["货值"]) else 0.0, 6),
        ]
        for _, row in frame.iterrows()
    ]
    calculator = {
        "materials": materials,
        "cities": cities,
        "warehouses": warehouses,
        "orders": orders,
        "materialMeta": material_meta,
        "rows": calculator_rows,
        "latestMonth": int(frame["月份"].max()),
    }
    (source_dir / "material_freight_calculator_data.js").write_text(
        js_assignment("FREIGHT_CALCULATOR_SOURCE", calculator), encoding="utf-8"
    )
    return {"orders": len(order_details), "materialRows": len(material_rows), "calculatorRows": len(calculator_rows)}


def build_after_sales(source_dir: Path, after_sales_path: Path) -> dict[str, object]:
    frame = pd.read_excel(after_sales_path, sheet_name="售后跟进", header=1)
    required = {"是否解决", "填写时间", "您在工作中遇到问题类型："}
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError("售后跟进缺少字段：" + "、".join(sorted(missing)))
    frame = frame.copy()
    frame["填写时间"] = pd.to_datetime(frame["填写时间"], errors="coerce")
    frame = frame.loc[frame["填写时间"].notna()].copy()
    if frame.empty:
        raise ValueError("售后跟进没有有效的填写时间")

    years = Counter(frame["填写时间"].dt.year.astype(int).tolist())
    year = max(years, key=lambda item: (years[item], item))
    frame = frame.loc[frame["填写时间"].dt.year.eq(year)].copy()
    categories = ["设备问题", "物料问题", "稽核问题", "其他问题"]
    after_sales = [0] * 12
    totals = [[0] * len(categories) for _ in range(12)]
    resolved = [[0] * len(categories) for _ in range(12)]
    for _, row in frame.iterrows():
        month_index = int(row["填写时间"].month) - 1
        after_sales[month_index] += 1
        issue_text = clean_text(row["您在工作中遇到问题类型："])
        is_resolved = clean_text(row["是否解决"]) in {"是", "已解决", "解决"}
        for category_index, category in enumerate(categories):
            if category in issue_text:
                totals[month_index][category_index] += 1
                if is_resolved:
                    resolved[month_index][category_index] += 1

    html_path = source_dir / "主看版.html"
    html = html_path.read_text(encoding="utf-8")
    html = html.replace(
        "数据源：物料及设备报表集合.xlsx · 物料开发进度跟进表",
        "数据源：物料开发进度跟进表.xlsx",
    )
    replacements = [
        (r"year:\s*\d+", f"year:{year}"),
        (r"afterSales:\s*\[[^\]]*\]", "afterSales:" + json.dumps(after_sales, ensure_ascii=False, separators=(",", ":"))),
    ]
    for pattern, replacement in replacements:
        html, count = re.subn(pattern, replacement, html, count=1, flags=re.S)
        if count != 1:
            raise ValueError(f"主看板售后数据块匹配失败：{pattern}")
    html_path.write_text(html, encoding="utf-8", newline="\n")
    return {"year": year, "records": len(frame), "months": after_sales}


DEVICE_BRAND_KEYWORDS = ("皇家小虎", "酷福", "美的", "智虎", "合马", "礼悦家", "海尔", "星星", "长才", "九达", "爱雪")


def build_material_weekly_outbound(source_dir: Path, inventory_path: Path) -> dict[str, object]:
    """从出入库流水 sheet 汇总本周/本月物料出库总量（周次=K列，数量=B列，仅出库）。"""
    frame = pd.read_excel(inventory_path, sheet_name="出入库流水", engine="openpyxl")
    required = {"周次", "数量", "出入库", "日期"}
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError("出入库流水 sheet 缺少字段：" + "、".join(sorted(missing)))
    frame = frame.copy()
    frame["数量"] = pd.to_numeric(frame["数量"], errors="coerce").fillna(0.0)
    frame["周次"] = frame["周次"].astype(str).str.strip()
    frame["出入库"] = frame["出入库"].astype(str).str.strip()
    frame["日期值"] = pd.to_datetime(frame["日期"], errors="coerce")
    outbound = frame.loc[frame["出入库"] == "出库"]

    def week_total(week: str) -> float:
        subset = outbound.loc[outbound["周次"] == week]
        return round(float(subset["数量"].sum()))

    month_qty = 0.0
    month_label = ""
    valid_dates = outbound["日期值"].dropna()
    if len(valid_dates):
        current_month = int(valid_dates.max().month)
        month_subset = outbound.loc[outbound["日期值"].dt.month == current_month]
        month_qty = round(float(month_subset["数量"].sum()))
        month_label = f"{current_month}月"

    data = {
        "thisWeek": week_total("本周"),
        "lastWeek": week_total("上周"),
        "monthQty": month_qty,
        "monthLabel": month_label,
        "weekLabel": "本周",
    }
    (source_dir / "material_weekly_outbound_data.js").write_text(
        js_assignment("MATERIAL_WEEKLY_OUTBOUND_DATA", data, spaced=True), encoding="utf-8"
    )
    return data


def extract_device_brand(name: str) -> str:
    """从设备名称中提取品牌，长关键词优先匹配。"""
    for keyword in DEVICE_BRAND_KEYWORDS:
        if keyword in name:
            return keyword
    return ""


def build_device_outbound_detail(
    frame: pd.DataFrame,
    current_month: int,
    this_week_range: tuple[pd.Timestamp, pd.Timestamp] | None = None,
    last_week_range: tuple[pd.Timestamp, pd.Timestamp] | None = None,
) -> dict[str, object]:
    """按弹窗表格口径生成省区/鸣忙设备发货明细：本月量金额、本周量金额及环比差值。

    本周/上周过滤规则：
    - 若传入 this_week_range / last_week_range（首尾日期闭区间），按日期范围切片；
    - 否则回退到「周次」列文本等于"本周"/"上周"的旧规则（兼容历史数据）。
    """
    data = frame.copy()
    data["数量"] = pd.to_numeric(data["数量"], errors="coerce").fillna(0.0)
    data["设备单价"] = pd.to_numeric(data.get("设备单价"), errors="coerce").fillna(0.0)
    data["金额"] = data["数量"] * data["设备单价"]
    data["月份值"] = pd.to_numeric(data.get("月份"), errors="coerce")
    data["大类"] = data["销售部门清洗-大类"].map(lambda v: "鸣忙" if clean_text(v) == "鸣忙" else "省区")
    data["设备"] = data["设备名称清洗"].map(lambda v: clean_text(v))
    data["下单时间"] = pd.to_datetime(data.get("下单日期"), errors="coerce")

    if this_week_range and last_week_range:
        this_start, this_end = pd.Timestamp(this_week_range[0]), pd.Timestamp(this_week_range[1])
        last_start, last_end = pd.Timestamp(last_week_range[0]), pd.Timestamp(last_week_range[1])
        this_week = data.loc[(data["下单时间"] >= this_start) & (data["下单时间"] <= this_end)]
        last_week = data.loc[(data["下单时间"] >= last_start) & (data["下单时间"] <= last_end)]
    else:
        def week_slice(week: str) -> pd.DataFrame:
            return data.loc[data["周次"].astype(str).str.strip() == week]
        this_week, last_week = week_slice("本周"), week_slice("上周")
    month_data = data.loc[data["月份值"] == current_month]

    groups: list[dict[str, object]] = []
    for key, group_value, title in (("province", "省区", "省区"), ("mingmang", "鸣忙", "鸣忙")):
        g_all = data.loc[data["大类"] == group_value]
        g_month = month_data.loc[month_data["大类"] == group_value]
        g_this = this_week.loc[this_week["大类"] == group_value]
        g_last = last_week.loc[last_week["大类"] == group_value]
        names = sorted(
            set(g_month["设备"]) | set(g_this["设备"]) | set(g_last["设备"]),
            key=lambda n: (DEVICE_BRAND_KEYWORDS.index(extract_device_brand(n)) if extract_device_brand(n) in DEVICE_BRAND_KEYWORDS else len(DEVICE_BRAND_KEYWORDS), n),
        )
        rows = []
        for name in names:
            if not name:
                continue
            month_qty = float(g_month.loc[g_month["设备"] == name, "数量"].sum())
            month_amount = float(g_month.loc[g_month["设备"] == name, "金额"].sum())
            week_qty = float(g_this.loc[g_this["设备"] == name, "数量"].sum())
            week_amount = float(g_this.loc[g_this["设备"] == name, "金额"].sum())
            last_qty = float(g_last.loc[g_last["设备"] == name, "数量"].sum())
            last_amount = float(g_last.loc[g_last["设备"] == name, "金额"].sum())
            rows.append(
                {
                    "brand": extract_device_brand(name) or "-",
                    "name": name,
                    "monthQty": round(month_qty),
                    "monthAmount": round(month_amount),
                    "weekQty": round(week_qty),
                    "qtyDelta": round(week_qty - last_qty),
                    "weekAmount": round(week_amount),
                    "amountDelta": round(week_amount - last_amount),
                    "note": "",
                }
            )
        total = {
            "monthQty": round(float(g_month["数量"].sum())),
            "monthAmount": round(float(g_month["金额"].sum())),
            "weekQty": round(float(g_this["数量"].sum())),
            "qtyDelta": round(float(g_this["数量"].sum() - g_last["数量"].sum())),
            "weekAmount": round(float(g_this["金额"].sum())),
            "amountDelta": round(float(g_this["金额"].sum() - g_last["金额"].sum())),
        }
        groups.append({"key": key, "title": title, "rows": rows, "total": total})
    return {"month": int(current_month), "groups": groups}


def read_device_week_range(device_path: Path) -> tuple[pd.Timestamp | None, pd.Timestamp | None, pd.Timestamp | None, pd.Timestamp | None]:
    """从关键字段汇总 sheet 读取 P2（上周末日期）/ Q2（本周末日期），推导本周/上周日期范围。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (None, None, None, None)
    try:
        wb = load_workbook(device_path, data_only=True, keep_vba=True, read_only=True)
        ws = wb["关键字段汇总"]
        o2 = ws["O2"].value
        p2 = ws["P2"].value
        q2 = ws["Q2"].value
        wb.close()
    except Exception:
        return (None, None, None, None)
    def _to_ts(value):
        if value is None:
            return None
        try:
            return pd.Timestamp(value)
        except Exception:
            return None
    last_start, last_end, this_end = _to_ts(o2), _to_ts(p2), _to_ts(q2)
    if last_end is None or this_end is None:
        return (None, None, None, None)
    if last_start is None:
        last_start = last_end - pd.Timedelta(days=6)
    this_start = last_end + pd.Timedelta(days=1)
    return (this_start, this_end, last_start, last_end)


def format_week_label(start: pd.Timestamp | None, end: pd.Timestamp | None) -> str:
    if start is None or end is None:
        return ""
    return f"{start.strftime('%m.%d')} ~ {end.strftime('%m.%d')}"


def build_device_weekly_summary(source_dir: Path, device_path: Path) -> dict[str, object]:
    """从关键字段汇总 sheet 汇总本周线下/鸣忙设备出库数量，并生成弹窗明细数据。

    弹窗明细预生成 detailByMonth（按月键），供前端按顶端月份下拉切换；周次范围基于 P2/Q2。
    """
    frame = pd.read_excel(device_path, sheet_name="关键字段汇总")
    required = {"周次", "销售部门清洗-大类", "数量", "设备名称清洗", "设备单价", "月份"}
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError("关键字段汇总 sheet 缺少字段：" + "、".join(sorted(missing)))

    this_start, this_end, last_start, last_end = read_device_week_range(device_path)
    week_frame = frame.copy()
    week_frame["数量"] = pd.to_numeric(week_frame["数量"], errors="coerce")
    week_frame = week_frame.loc[week_frame["数量"].notna()].copy()
    week_frame["下单时间"] = pd.to_datetime(week_frame.get("下单日期"), errors="coerce")

    if this_start is not None and this_end is not None:
        week_frame = week_frame.loc[
            (week_frame["下单时间"] >= this_start) & (week_frame["下单时间"] <= this_end)
        ]

    def classify(value: object) -> str:
        return "鸣忙" if clean_text(value) == "鸣忙" else "线下"

    week_frame["类型"] = week_frame["销售部门清洗-大类"].map(classify)
    summary = week_frame.groupby("类型")["数量"].sum().to_dict()

    month_values = pd.to_numeric(frame.get("月份"), errors="coerce").dropna()
    month_list = sorted({int(m) for m in month_values if pd.notna(m)})

    # 预生成各月份弹窗明细
    detail_by_month: dict[str, dict[str, object]] = {}
    for month_value in month_list:
        month_detail = build_device_outbound_detail(
            frame,
            month_value,
            this_week_range=(this_start, this_end) if this_start is not None else None,
            last_week_range=(last_start, last_end) if last_start is not None else None,
        )
        detail_by_month[str(month_value)] = month_detail

    latest_month = month_list[-1] if month_list else 1
    data = {
        "offline": round(float(summary.get("线下", 0.0)), 2),
        "mingmang": round(float(summary.get("鸣忙", 0.0)), 2),
        "weekLabel": "本周",
        "weekRange": {
            "thisStart": this_start.strftime("%Y-%m-%d") if this_start is not None else "",
            "thisEnd": this_end.strftime("%Y-%m-%d") if this_end is not None else "",
            "lastStart": last_start.strftime("%Y-%m-%d") if last_start is not None else "",
            "lastEnd": last_end.strftime("%Y-%m-%d") if last_end is not None else "",
            "thisLabel": format_week_label(this_start, this_end),
            "lastLabel": format_week_label(last_start, last_end),
        },
        "monthList": month_list,
        "latestMonth": latest_month,
        "detailByMonth": detail_by_month,
    }
    (source_dir / "device_weekly_outbound_data.js").write_text(
        js_assignment("DEVICE_WEEKLY_OUTBOUND_DATA", data, spaced=True), encoding="utf-8"
    )
    return {"summary": data, "rows": len(week_frame)}


def update(source_dir: Path) -> dict[str, object]:
    paths = {
        "inventory": source_dir / INVENTORY_FILE,
        "device": resolve_device_path(source_dir),
        "freight": source_dir / FREIGHT_FILE,
        "afterSales": source_dir / AFTER_SALES_FILE,
        "development": source_dir / DEVELOPMENT_FILE,
    }
    missing = [path.name for path in paths.values() if not path.exists()]
    if missing:
        raise FileNotFoundError("缺少独立源数据：" + "、".join(missing))

    legacy = load_legacy_updater(source_dir)
    original_read_excel = legacy.pd.read_excel

    device_frame = None

    def routed_read_excel(io, *args, **kwargs):
        nonlocal device_frame
        sheet = kwargs.get("sheet_name")
        if sheet is None and args:
            sheet = args[0]
        if sheet == "设备分析":
            if device_frame is None:
                device_frame, _ = load_device_workbook(paths["device"])
            return device_frame.copy()
        route = {"物料运费分析": paths["freight"], "物料开发进度跟进表": paths["development"]}
        frame = original_read_excel(route.get(sheet, io), *args, **kwargs)
        return normalize_freight_columns(frame) if sheet == "物料运费分析" else frame

    legacy.INVENTORY_SOURCE = paths["inventory"]
    legacy.MASTER_SOURCE = paths["freight"]
    legacy.pd.read_excel = routed_read_excel
    try:
        inventory = legacy.update_inventory()
        freight = legacy.update_freight()
        development = legacy.update_gantt()
    finally:
        legacy.pd.read_excel = original_read_excel

    device = build_device_outbound_file(source_dir, paths["device"])
    freight_details = build_freight_detail_files(source_dir, paths["freight"])
    after_sales = build_after_sales(source_dir, paths["afterSales"])
    weekly_device = build_device_weekly_summary(source_dir, paths["device"])
    weekly_material = build_material_weekly_outbound(source_dir, paths["inventory"])

    # 生成带版本号的副本，避免本地 file:// 浏览器对原文件名缓存顽固导致拿不到最新数据
    import shutil
    for src_name, dst_name in (
        ("device_weekly_outbound_data.js", "device_weekly_outbound_data.v2.js"),
        ("material_weekly_outbound_data.js", "material_weekly_outbound_data.v2.js"),
    ):
        src_path = source_dir / src_name
        dst_path = source_dir / dst_name
        if src_path.exists():
            shutil.copyfile(src_path, dst_path)
    result = {
        "sources": {key: path.name for key, path in paths.items()},
        "inventory": inventory,
        "device": device,
        "freight": freight,
        "freightDetails": freight_details,
        "development": development,
        "afterSales": after_sales,
        "weeklyDevice": weekly_device,
        "weeklyMaterial": weekly_material,
    }
    (source_dir / "最近一次更新记录.txt").write_text(
        json.dumps({"status": "success", **result}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="从独立 Excel 数据源同步更新三个物料看板")
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--device-only", action="store_true", help="仅更新进销存看板的设备数据文件")
    args = parser.parse_args()
    if args.device_only:
        result = build_device_outbound_file(args.source_dir.resolve(), resolve_device_path(args.source_dir.resolve()))
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return
    result = update(args.source_dir.resolve())
    summary = {
        "sources": result["sources"],
        "inventory": {
            "dataDate": result["inventory"]["dataDate"],
            "materials": result["inventory"]["materials"],
            "flows": result["inventory"]["flows"],
        },
        "device": result["device"],
        "freight": result["freight"],
        "freightDetails": result["freightDetails"],
        "development": result["development"],
        "afterSales": result["afterSales"],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
