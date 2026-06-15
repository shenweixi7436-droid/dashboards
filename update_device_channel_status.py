from pathlib import Path
import json
from openpyxl import load_workbook


SOURCE = Path("C:/Users/shenw/Desktop/\u65b0\u5efa\u6587\u4ef6\u5939/\u5e02\u573a\u7a3d\u6838\u90e8\u91cd\u70b9\u5de5\u4f5c.xlsx")
SHEET = "\u667a\u80fd\u8bbe\u5907\u53f0\u8d26\u6c47\u603b"
OUTPUT = Path(__file__).resolve().parent / "assets" / "data" / "device-channel-status.js"


def main():
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    sheet = workbook[SHEET]

    # The chart data range is M4:O6. Device names are fixed to avoid workbook
    # cached text encoding issues, while values still come from the workbook.
    device_names = ["\u4fdd\u6e29\u67dc", "\u70e4\u80a0\u673a", "\u51b0\u67dc"]
    items = []
    for index, row in enumerate(range(4, 7)):
        volume = sheet.cell(row, 14).value or 0
        rate = sheet.cell(row, 15).value or 0
        items.append({
            "name": device_names[index],
            "volume": int(float(volume)),
            "rate": float(rate),
        })

    payload = {
        "source": "\u5e02\u573a\u7a3d\u6838\u90e8\u91cd\u70b9\u5de5\u4f5c.xlsx / \u667a\u80fd\u8bbe\u5907\u53f0\u8d26\u6c47\u603b!M4:O6",
        "items": items,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        "window.DEVICE_CHANNEL_STATUS = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Updated: {OUTPUT}")


if __name__ == "__main__":
    main()
