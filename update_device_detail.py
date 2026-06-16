from pathlib import Path
import json
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = Path("C:/Users/shenw/Desktop/\u770b\u677f/\u5e02\u573a\u7a3d\u6838\u90e8\u91cd\u70b9\u5de5\u4f5c.xlsx")
OUT = ROOT / "assets" / "data" / "device-detail.js"
SHEET = "\u667a\u80fd\u8bbe\u5907\u53f0\u8d26\u6c47\u603b"


def norm(value):
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 4)
    return value


def display(value, is_rate=False):
    value = norm(value)
    if value == "":
        return "-"
    if is_rate and isinstance(value, (int, float)):
        return f"{value * 100:.1f}%"
    if isinstance(value, (int, float)):
        if abs(value - int(value)) < 0.00001:
            return f"{int(value):,}"
        return f"{value:.1f}"
    return str(value)


def read_table(ws, name, header_row, start_row, end_row, start_col=1, end_col=8):
    headers = [display(ws.cell(header_row, col).value) for col in range(start_col, end_col + 1)]
    rows = []
    for row_idx in range(start_row, end_row + 1):
        raw = [ws.cell(row_idx, col).value for col in range(start_col, end_col + 1)]
        if not any(value not in (None, "") for value in raw):
            continue
        rows.append(
            [
                display(value, is_rate=(idx == len(raw) - 1 and isinstance(value, (int, float)) and value <= 1))
                for idx, value in enumerate(raw)
            ]
        )
    return {"name": name, "headers": headers, "rows": rows}


wb = load_workbook(SOURCE, data_only=True, read_only=True)
ws = wb[SHEET]

sections = [
    read_table(ws, "保温柜", 3, 4, 7, 1, 8),
    read_table(ws, "烤肠机", 9, 10, 13, 1, 8),
    read_table(ws, "冰柜", 15, 16, 18, 1, 7),
]

summary = []
for row in range(4, 7):
    name = display(ws.cell(row, 13).value)
    volume = ws.cell(row, 14).value or 0
    rate = ws.cell(row, 15).value or 0
    if name and name != "-":
        summary.append({"name": name, "volume": int(volume), "rate": float(rate)})

payload = {
    "month": 6,
    "source": "市场稽核部重点工作.xlsx / 智能设备台账汇总",
    "summary": summary,
    "sections": sections,
}

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(
    "window.DEVICE_DETAIL = "
    + json.dumps(payload, ensure_ascii=False, indent=2)
    + ";\n",
    encoding="utf-8",
)
print(f"已生成: {OUT}")
print("设备明细表数量:", len(sections))
