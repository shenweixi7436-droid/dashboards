from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import json
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = Path("C:/Users/shenw/Desktop/看板/市场稽核部重点工作.xlsx")
DATA_DIR = ROOT / "assets" / "data"
MARKET_ORDER_IMAGE_DIR = ROOT / "assets" / "images" / "market-order"

PROMO_AUDIT_SHEET = "推广促销稽核"
PROMO_PLAN_SHEET = "推广促销计划"
APPROVAL_SHEET = "线上审批流程稽核明细"
DEVICE_SHEET = "智能设备台账汇总"
MARKET_ORDER_CASE_SHEET = "市场秩序治理-窜货案件数"
MARKET_ORDER_CUSTOMER_SHEET = "市场秩序治理-涉及客户明细"
GIFT_ACTIVITY_SHEET = "赠品稽核-活动"
GIFT_SAMPLE_SHEET = "赠品稽核-样品"
DEVICE_BAN_SHEET = "禁网行动"
STORE_AUDIT_SHEET = "鸣忙门店专项稽核"


def norm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def number(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def as_rate(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        v = float(value)
    except (TypeError, ValueError):
        return 0.0
    return v if v <= 1 else v / 100


def month_label(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return f"{value.month}月"
    text = str(value).strip()
    m = re.search(r"(\d{1,2})\s*月", text)
    if m:
        return f"{int(m.group(1))}月"
    try:
        dt = datetime.fromisoformat(text[:10])
        return f"{dt.month}月"
    except ValueError:
        pass
    if text.isdigit() and 1 <= int(text) <= 12:
        return f"{int(text)}月"
    return ""


def month_key(month: str) -> int:
    m = re.search(r"\d+", str(month))
    return int(m.group(0)) if m else 99


def result_label(value) -> str:
    text = norm(value)
    return "合格" if text in {"是", "合格"} else "不合格"


def derive_month(row_values, preferred_indexes):
    for idx in preferred_indexes:
        if idx < len(row_values):
            month = month_label(row_values[idx])
            if month:
                return month
    return ""


def read_headers(ws, row, cols, fallback):
    headers = []
    for col, fb in zip(cols, fallback):
        headers.append(norm(ws.cell(row, col).value) or fb)
    return headers


def build_promo_plan(ws):
    rows_by_month = defaultdict(dict)
    months = set()
    max_col = ws.max_column
    month_starts = [col for col in range(1, max_col + 1) if norm(ws.cell(1, col).value).startswith("月份")]
    for start_col in month_starts:
        for row in range(2, ws.max_row + 1):
            raw = [ws.cell(row, start_col + offset).value for offset in range(7)]
            month = month_label(raw[0])
            province = norm(raw[2])
            target = number(raw[4])
            if not month or not province:
                continue
            months.add(month)
            rows_by_month[month][province] = rows_by_month[month].get(province, 0) + target
    return rows_by_month, months


def build_promo(wb):
    ws_audit = wb[PROMO_AUDIT_SHEET]
    ws_plan = wb[PROMO_PLAN_SHEET]
    plan_by_month, plan_months = build_promo_plan(ws_plan)

    detail_cols = [2, 3, 4, 6, 14]
    headers = read_headers(
        ws_audit,
        3,
        detail_cols,
        ["省份", "市区", "城市经理", "活动门店名称", "稽核结果"],
    )
    rows_by_month = defaultdict(list)
    audit_by_month_province = defaultdict(Counter)
    totals_by_month = defaultdict(lambda: {"total": 0, "qualified": 0, "unqualified": 0})
    months = set(plan_months)

    for row_idx, raw_row in enumerate(ws_audit.iter_rows(min_row=4, values_only=True), 4):
        values = [norm(raw_row[col - 1] if len(raw_row) >= col else "") for col in detail_cols]
        result = norm(raw_row[13] if len(raw_row) >= 14 else "")
        if not any(values) and not result:
            continue
        # 推广促销稽核按表内月份归属；没有月份的行不归入任何月份，
        # 避免切换到其他月份时被当前系统月份误带出数据。
        month = derive_month(raw_row, [0, 11, 8])
        if not month:
            continue
        months.add(month)
        # Column B is the provincial manager; column C is the province name.
        province = norm(raw_row[2] if len(raw_row) >= 3 else "")
        if province:
            audit_by_month_province[month][province] += 1
        label = result_label(result)
        totals_by_month[month]["total"] += 1
        if label == "合格":
            totals_by_month[month]["qualified"] += 1
        else:
            totals_by_month[month]["unqualified"] += 1
        rows_by_month[month].append({"row": row_idx, "values": values, "result": label})

    plan_payload = {}
    detail_payload = {}
    for month in sorted(months, key=month_key):
        provinces = sorted(set(plan_by_month.get(month, {})) | set(audit_by_month_province.get(month, {})))
        rows = []
        for province in provinces:
            rows.append({
                "province": province,
                "plan": int(plan_by_month.get(month, {}).get(province, 0)),
                "audit": int(audit_by_month_province.get(month, {}).get(province, 0)),
            })
        plan_total = sum(r["plan"] for r in rows)
        audit_total = sum(r["audit"] for r in rows)
        plan_payload[month] = {
            "month": month,
            "planTotal": plan_total,
            "auditTotal": audit_total,
            "progress": round(audit_total / plan_total * 100, 1) if plan_total else 0,
            "rows": rows,
        }
        total = totals_by_month[month]["total"]
        detail_payload[month] = {
            "month": month_key(month),
            "source": "市场稽核部重点工作.xlsx / 推广促销稽核",
            "headers": headers,
            "total": total,
            "qualified": totals_by_month[month]["qualified"],
            "unqualified": totals_by_month[month]["unqualified"],
            "rows": rows_by_month.get(month, []),
        }
    return plan_payload, detail_payload, months


def build_approval(wb):
    ws = wb[APPROVAL_SHEET]
    # Current sheet layout (row 2):
    # A 月, B 审核日期, C 流程发起时间, D 审核人, E 费用类型,
    # F 签呈号, G 省区, H 客户名称, I 发起人, J 是否合格,
    # K 问题类型, L 着装不合格类型.
    # The previous offsets were one column too far left, so the dashboard
    # mistakenly treated 签呈号 as 省区 and 发起人 as 是否合格.
    detail_cols = [3, 5, 6, 7, 8, 9, 11, 12]
    headers = read_headers(
        ws,
        2,
        detail_cols,
        ["流程发起时间", "费用类型", "签呈号", "省区", "客户名称", "发起人", "问题类型", "着装不合格类型"],
    )
    stats = defaultdict(lambda: {"total": 0, "qualified": 0, "unqualified": 0, "issues": Counter(), "dress": Counter(), "province": Counter(), "rows": []})
    months = set()
    for row_idx, raw_row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
        values = [norm(raw_row[col - 1] if len(raw_row) >= col else "") for col in detail_cols]
        result = norm(raw_row[9] if len(raw_row) >= 10 else "")
        if not any(values) and not result:
            continue
        # 线上审批流程稽核明细按 B 列拆分月份；无月份的行不归入任何月份。
        month = derive_month(raw_row, [1])
        if not month:
            continue
        months.add(month)
        bucket = stats[month]
        bucket["total"] += 1
        if result == "是":
            bucket["qualified"] += 1
        else:
            bucket["unqualified"] += 1
            issue = values[6]
            dress_issue = values[7] if len(values) > 7 else ""
            province = values[3]
            if issue:
                bucket["issues"][issue] += 1
            if dress_issue:
                bucket["dress"][dress_issue] += 1
            if province:
                bucket["province"][province] += 1
            bucket["rows"].append({"row": row_idx, "values": values, "result": "不合格"})

    pies_payload = {}
    detail_payload = {}
    for month in sorted(months, key=month_key):
        bucket = stats[month]
        total = bucket["total"]
        issues = [{"name": k, "value": v} for k, v in bucket["issues"].most_common(3)]
        dress_issues = [{"name": k, "value": v} for k, v in bucket["dress"].most_common()]
        province_issues = [{"province": k, "value": v} for k, v in bucket["province"].most_common() if v > 0]
        pies_payload[month] = {
            "month": month,
            "total": total,
            "qualified": bucket["qualified"],
            "unqualified": bucket["unqualified"],
            "rate": round(bucket["qualified"] / total * 100, 1) if total else 0,
            "issues": issues,
            "dressIssues": dress_issues,
            "provinceIssues": province_issues,
        }
        detail_payload[month] = {
            "month": month_key(month),
            "source": "市场稽核部重点工作.xlsx / 线上审批流程稽核明细",
            "headers": headers,
            "total": total,
            "qualified": bucket["qualified"],
            "unqualified": bucket["unqualified"],
            "rows": bucket["rows"],
            "dressIssues": dress_issues,
            "provinceIssues": province_issues,
        }
    return pies_payload, detail_payload, months


def build_device(wb, months):
    ws = wb[DEVICE_SHEET]

    def display(value, is_rate=False):
        value = "" if value is None else value
        if value == "":
            return "-"
        if is_rate and isinstance(value, (int, float)):
            return f"{value * 100:.1f}%"
        if isinstance(value, (int, float)):
            return f"{int(value):,}" if abs(value - int(value)) < 0.00001 else f"{value:.1f}"
        return str(value).strip()

    def row_values(row):
        return [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]

    def find_label_row(label, start_row=1):
        for row in range(start_row, ws.max_row + 1):
            values = [norm(value) for value in row_values(row)]
            if label in values and any("设备投放数量统计" in value for value in values):
                return row
        return None

    def find_row_with_labels(start_row, labels, search_rows=6):
        end_row = min(ws.max_row, start_row + search_rows)
        for row in range(start_row, end_row + 1):
            values = [norm(value) for value in row_values(row)]
            if all(label in values for label in labels):
                return row
        return None

    def find_value_row(start_row, label, search_rows=12):
        end_row = min(ws.max_row, start_row + search_rows)
        for row in range(start_row, end_row + 1):
            for col in range(1, ws.max_column + 1):
                if norm(ws.cell(row, col).value) == label:
                    return row, col
        return None, None

    def header_map(row):
        return {
            norm(ws.cell(row, col).value): col
            for col in range(1, ws.max_column + 1)
            if norm(ws.cell(row, col).value)
        }

    def read_table(name, header_row, start_row, end_row, start_col, end_col):
        headers = [display(ws.cell(header_row, col).value) for col in range(start_col, end_col + 1)]
        rows = []
        for row_idx in range(start_row, end_row + 1):
            raw = [ws.cell(row_idx, col).value for col in range(start_col, end_col + 1)]
            if not any(value not in (None, "") for value in raw):
                continue
            rows.append([
                display(value, is_rate=(idx == len(raw) - 1 and isinstance(value, (int, float)) and value <= 1))
                for idx, value in enumerate(raw)
            ])
        return {"name": name, "headers": headers, "rows": rows}

    def read_standard_device(name):
        title_row = find_label_row(name)
        if not title_row:
            return None
        headers_row = find_row_with_labels(title_row + 1, ["渠道", "已投放"])
        if not headers_row:
            return None
        headers = header_map(headers_row)
        total_row, channel_col = find_value_row(headers_row + 1, "总计")
        if not total_row or not channel_col:
            return None
        volume_col = headers.get("已投放")
        active_col = headers.get("开机数量") or headers.get("开机")
        end_col = max(col for col in (volume_col, active_col, channel_col) if col)
        volume = number(ws.cell(total_row, volume_col).value) if volume_col else 0
        active = number(ws.cell(total_row, active_col).value) if active_col else 0
        return {
            "name": name,
            "volume": volume,
            "active": active,
            # 统一按开机数量 / 已投放计算开机率，不再使用台效达标率。
            "rate": round(active / volume, 4) if volume else 0,
            "section": read_table(
                name,
                headers_row,
                headers_row + 1,
                total_row,
                channel_col,
                end_col,
            ),
        }

    standard_devices = []
    for device_name in ("保温柜", "烤肠机", "星星冰柜"):
        device = read_standard_device(device_name)
        if device:
            standard_devices.append(device)

    items = [
        {
            "name": device["name"],
            "volume": device["volume"],
            "active": device["active"],
            "rate": device["rate"],
        }
        for device in standard_devices
    ]

    haier = {
        "name": "海尔冰柜",
        "volume": 0,
        "rate": 0,
        "ledger": 0,
        "system": 0,
        "difference": 0,
        "active": 0,
        "inactive": 0,
        "inStock": 0,
        "atStore": 0,
        "channels": [],
    }
    haier_section = None
    haier_title_row = find_label_row("海尔冰柜")
    if haier_title_row:
        subheader_row = find_row_with_labels(
            haier_title_row + 1,
            ["台账数据", "系统数据", "差异", "总数", "在库", "门店"],
            search_rows=8,
        )
        if subheader_row:
            headers = header_map(subheader_row)
            active_col = None
            for row in range(haier_title_row + 1, subheader_row + 1):
                for col in range(1, ws.max_column + 1):
                    if norm(ws.cell(row, col).value) == "开机":
                        active_col = col
                        break
                if active_col:
                    break
            total_row, channel_col = find_value_row(subheader_row + 1, "合计")
            if total_row and channel_col:
                columns = {
                    "ledger": headers.get("台账数据"),
                    "system": headers.get("系统数据"),
                    "difference": headers.get("差异"),
                    "active": active_col,
                    "inactive": headers.get("总数"),
                    "inStock": headers.get("在库"),
                    "atStore": headers.get("门店"),
                }

                def read_haier_row(row):
                    result = {"name": norm(ws.cell(row, channel_col).value)}
                    for key, col in columns.items():
                        result[key] = number(ws.cell(row, col).value) if col else 0
                    # 海尔开机率统一按开机数 / 系统数量计算。
                    result["rate"] = round(result["active"] / result["system"], 4) if result["system"] else 0
                    return result

                channel_rows = [
                    read_haier_row(row)
                    for row in range(subheader_row + 1, total_row)
                    if norm(ws.cell(row, channel_col).value)
                ]
                total = read_haier_row(total_row)
                haier.update(total)
                haier["name"] = "海尔冰柜"
                haier["volume"] = total["ledger"]
                haier["channels"] = channel_rows
                end_col = max(col for col in columns.values() if col)
                haier_headers = ["渠道", "台账数据", "系统数据", "差异", "开机", "未开机总数", "在库", "门店"]
                haier_rows = []
                for row in channel_rows + [total]:
                    haier_rows.append([
                        row["name"],
                        f'{row["ledger"]:,}',
                        f'{row["system"]:,}',
                        f'{row["difference"]:,}',
                        f'{row["active"]:,}',
                        f'{row["inactive"]:,}',
                        f'{row["inStock"]:,}',
                        f'{row["atStore"]:,}',
                    ])
                haier_section = {"name": "海尔冰柜", "headers": haier_headers, "rows": haier_rows}

    overview = items + [{
        "name": "海尔冰柜",
        "volume": haier["volume"],
        "active": haier["active"],
        "rate": haier["rate"],
    }]
    detail = {
        "month": 6,
        "source": "市场稽核部重点工作.xlsx / 智能设备台账汇总",
        "summary": items,
        "sections": [device["section"] for device in standard_devices]
        + ([haier_section] if haier_section else []),
        "popup": {
            "source": "市场稽核部重点工作.xlsx / 智能设备台账汇总（按设备名称与表头识别）",
            "overview": overview,
            "ranking": items,
            "haier": haier,
        },
    }
    status = {
        "source": "市场稽核部重点工作.xlsx / 智能设备台账汇总（按设备名称与表头识别）",
        "items": items,
    }
    return {month: status for month in months}, {month: detail for month in months}


def build_market_order(wb):
    ws_cases = wb[MARKET_ORDER_CASE_SHEET]
    ws_customers = wb[MARKET_ORDER_CUSTOMER_SHEET]
    monthly = defaultdict(lambda: {
        "cases": set(),
        "customers": set(),
        "provinceCases": defaultdict(set),
        "customerRows": defaultdict(int),
        "locked": set(),
        "punish": set(),
        "internal": set(),
        "unverified": set(),
        "others": {},
        "caseDetails": {},
        "caseDetailOrder": [],
    })
    months = set()

    def sheet_headers(ws):
        return {norm(ws.cell(1, col).value): col - 1 for col in range(1, ws.max_column + 1)}

    case_headers = sheet_headers(ws_cases)

    def cell_value(row, header, fallback_index=None):
        idx = case_headers.get(header)
        if idx is None:
            idx = fallback_index
        if idx is None or idx >= len(row):
            return ""
        return norm(row[idx])

    def is_dispimg(value):
        return bool(re.search(r"(?:_xlfn\.)?DISPIMG\s*\(", norm(value), re.IGNORECASE))

    def normalize_verified(value):
        text = re.sub(r"\s+", "", norm(value))
        if "未查实" in text:
            return "未查实"
        if "已查实" in text:
            return "已查实"
        return text

    def build_image_map(ws):
        image_map = defaultdict(list)
        images = getattr(ws, "_images", [])
        if not images:
            return image_map
        MARKET_ORDER_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        for idx, img in enumerate(images, start=1):
            anchor = getattr(img, "anchor", None)
            if not hasattr(anchor, "_from"):
                continue
            row_no = anchor._from.row + 1
            col_no = anchor._from.col + 1
            fmt = (getattr(img, "format", None) or "png").lower()
            ext = "jpg" if fmt == "jpeg" else fmt
            filename = f"market-order-r{row_no}-c{col_no}-{idx}.{ext}"
            target = MARKET_ORDER_IMAGE_DIR / filename
            try:
                target.write_bytes(img._data())
                image_map[(row_no, col_no)].append(f"assets/images/market-order/{filename}")
            except Exception:
                continue
        return image_map

    case_image_map = build_image_map(ws_cases)
    smuggler_col = case_headers.get("窜货方")
    penalty_col = case_headers.get("处罚通告清单")

    def is_2026(value):
        text = norm(value)
        if isinstance(value, datetime):
            return value.year == 2026
        m = re.search(r"20\d{2}", text)
        return bool(m and int(m.group(0)) == 2026)

    for row_no, row in enumerate(ws_cases.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not is_2026(row[0] if len(row) > 0 else ""):
            continue
        month = month_label(row[1] if len(row) > 1 else "")
        seq = norm(row[2] if len(row) > 2 else "")
        if not month or not seq:
            continue
        months.add(month)
        bucket = monthly[month]
        bucket["cases"].add(seq)
        province = norm(row[4] if len(row) > 4 else "")
        city = norm(row[5] if len(row) > 5 else "")
        verified = normalize_verified(cell_value(row, "查实情况", 10))
        method = cell_value(row, "处理结果", 13)
        if province:
            bucket["provinceCases"][province].add(seq)
        if verified == "已查实":
            bucket["locked"].add(seq)
        if method in {"营销中心通报处罚", "省区通报处罚"}:
            bucket["punish"].add(seq)
        elif method in {"内部处理", "内部沟通处理"}:
            bucket["internal"].add(seq)
        elif method == "未查实":
            bucket["unverified"].add(seq)
        elif method:
            note = f"窜货{seq}{province}{city}{method}"
            bucket["others"][seq] = note
        penalty_images = case_image_map.get((row_no, penalty_col + 1), []) if penalty_col is not None else []
        smuggler_has_image = bool(case_image_map.get((row_no, smuggler_col + 1), [])) if smuggler_col is not None else False
        smuggler_value = cell_value(row, "窜货方", 12)
        penalty_notice = cell_value(row, "处罚通告清单", 14)
        case_detail = {
            "seq": seq,
            "auditDate": cell_value(row, "稽核日期", 3),
            "province": province,
            "city": city,
            "feedback": cell_value(row, "投诉反馈人", 6),
            "batch": cell_value(row, "投诉批次", 9),
            "verified": verified,
            "remark": cell_value(row, "备注", 11),
            "smuggler": "" if smuggler_has_image or is_dispimg(smuggler_value) else smuggler_value,
            "result": method,
            "penaltyNotice": "" if is_dispimg(penalty_notice) else penalty_notice,
            "penaltyImages": penalty_images,
        }
        if seq not in bucket["caseDetails"]:
            bucket["caseDetails"][seq] = case_detail
            bucket["caseDetailOrder"].append(seq)
        else:
            existing = bucket["caseDetails"][seq]
            for key, value in case_detail.items():
                if key == "penaltyImages":
                    existing[key] = list(dict.fromkeys((existing.get(key) or []) + (value or [])))
                elif key == "verified" and value:
                    # The last source row is the current verification status.
                    existing[key] = value
                elif key in {"remark", "result", "penaltyNotice"} and value:
                    existing[key] = value
                elif not existing.get(key) and value:
                    existing[key] = value

    for row in ws_customers.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        month = month_label(row[0] if len(row) > 0 else "")
        customer = norm(row[1] if len(row) > 1 else "")
        if not month or not customer:
            continue
        months.add(month)
        monthly[month]["customers"].add(customer)
        monthly[month]["customerRows"][customer] += 1

    payload = {}
    for month in sorted(months, key=month_key):
        bucket = monthly[month]
        case_details = []
        for seq in bucket["caseDetailOrder"]:
            if seq not in bucket["caseDetails"]:
                continue
            detail = dict(bucket["caseDetails"][seq])
            if seq in bucket["locked"]:
                detail["verified"] = "已查实"
            else:
                detail["verified"] = normalize_verified(detail.get("verified")) or "未查实"
            case_details.append(detail)
        province_rank = sorted(
            ({"name": name, "count": len(seq_set)} for name, seq_set in bucket["provinceCases"].items()),
            key=lambda item: (-item["count"], item["name"])
        )
        customer_rank = sorted(
            ({"name": name, "count": count} for name, count in bucket["customerRows"].items()),
            key=lambda item: (-item["count"], item["name"])
        )
        payload[month] = {
            "month": month,
            "source": "市场稽核部重点工作.xlsx / 市场秩序治理",
            "caseCount": len(bucket["cases"]),
            "customerCount": len(bucket["customers"]),
            "lockedCustomerCount": len(bucket["locked"]),
            "punishCount": len(bucket["punish"]),
            "internalCount": len(bucket["internal"]),
            "unverifiedCount": len(bucket["unverified"]),
            "otherNotes": list(bucket["others"].values()),
            "provinceRank": province_rank,
            "customerRank": customer_rank,
            "caseDetails": case_details,
        }
    return payload, months


def sum_numeric_column(ws, column: int, min_row: int = 1) -> float:
    """Sum every numeric cell in a source column, matching Excel's column SUM."""
    total = 0.0
    for (value,) in ws.iter_rows(
        min_row=min_row, min_col=column, max_col=column, values_only=True
    ):
        if isinstance(value, bool) or value in (None, ""):
            continue
        try:
            total += float(value)
        except (TypeError, ValueError):
            continue
    return total


def clean_total(value: float):
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded


def mom_rate(current: float, previous: float):
    if not previous:
        return 0.0 if not current else None
    return round((current - previous) / previous * 100, 1)


def ranked_amounts(ws, name_column: int, amount_column: int, limit: int = 5):
    totals = defaultdict(float)
    first_column = min(name_column, amount_column)
    for row in ws.iter_rows(
        min_row=5,
        min_col=first_column,
        max_col=max(name_column, amount_column),
        values_only=True,
    ):
        name = row[name_column - first_column]
        amount = row[amount_column - first_column]
        label = norm(name)
        if not label or not isinstance(amount, (int, float)) or isinstance(amount, bool):
            continue
        totals[label] += float(amount)
    ranked = sorted(totals.items(), key=lambda item: (-item[1], item[0]))[:limit]
    return [{"name": name, "amount": clean_total(amount)} for name, amount in ranked]


def metric_pair(ws, current_column: int, previous_column: int, min_row: int = 1):
    current = sum_numeric_column(ws, current_column, min_row=min_row)
    previous = sum_numeric_column(ws, previous_column, min_row=min_row)
    return current, previous


def build_gift_audit(wb):
    activity_ws = wb[GIFT_ACTIVITY_SHEET]
    sample_ws = wb[GIFT_SAMPLE_SHEET]
    activity_blocks = sorted(
        [
            {
                "month": month_label(activity_ws.cell(1, 30).value),
                "name": 17,
                "dealerCount": 21,
                "dealerAmount": 22,
                "consumerGiftCount": 23,
                "consumerGiftAmount": 24,
                "tasteCount": 25,
                "tasteAmount": 26,
                "totalCount": 29,
                "totalAmount": 30,
            },
            {
                "month": month_label(activity_ws.cell(1, 45).value),
                "name": 32,
                "dealerCount": 36,
                "dealerAmount": 37,
                "consumerGiftCount": 38,
                "consumerGiftAmount": 39,
                "tasteCount": 40,
                "tasteAmount": 41,
                "totalCount": 44,
                "totalAmount": 45,
            },
        ],
        key=lambda block: month_key(block["month"]),
        reverse=True,
    )
    sample_blocks = sorted(
        [
            {
                "month": month_label(sample_ws.cell(1, 27).value),
                "name": 20,
                "oldCount": 21,
                "oldAmount": 22,
                "newCount": 23,
                "newAmount": 24,
                "customerNew": 25,
                "totalCount": 26,
                "totalAmount": 27,
            },
            {
                "month": month_label(sample_ws.cell(1, 43).value),
                "name": 36,
                "oldCount": 38,
                "oldAmount": 39,
                "newCount": 40,
                "newAmount": 41,
                "customerNew": 42,
                "totalCount": 42,
                "totalAmount": 43,
            },
        ],
        key=lambda block: month_key(block["month"]),
        reverse=True,
    )
    activity_current, activity_previous = activity_blocks
    sample_current, sample_previous = sample_blocks
    if (
        activity_current["month"]
        and sample_current["month"]
        and activity_current["month"] != sample_current["month"]
    ):
        raise ValueError(
            f"赠品稽核月份不一致：活动 {activity_current['month']}，样品 {sample_current['month']}"
        )

    def category_payload(ws, current_block, previous_block, count_key, amount_key):
        current_count, previous_count = metric_pair(
            ws, current_block[count_key], previous_block[count_key], min_row=5
        )
        current_amount, previous_amount = metric_pair(
            ws, current_block[amount_key], previous_block[amount_key], min_row=5
        )
        return {
            "count": clean_total(current_count),
            "amount": clean_total(current_amount),
            "countMom": mom_rate(current_count, previous_count),
            "amountMom": mom_rate(current_amount, previous_amount),
        }

    activity_count, previous_activity_count = metric_pair(
        activity_ws,
        activity_current["totalCount"],
        activity_previous["totalCount"],
    )
    activity_amount, previous_activity_amount = metric_pair(
        activity_ws,
        activity_current["totalAmount"],
        activity_previous["totalAmount"],
    )
    sample_count, previous_sample_count = metric_pair(
        sample_ws, sample_current["totalCount"], sample_previous["totalCount"]
    )
    sample_amount, previous_sample_amount = metric_pair(
        sample_ws, sample_current["totalAmount"], sample_previous["totalAmount"]
    )
    # 客户新增严格按样品表 Y 列（当月客户开发数）汇总；历史月度区块未提供同口径列，环比不展示。
    customer_new_count = sum_numeric_column(sample_ws, sample_current["customerNew"], min_row=5)

    activity_rank_columns = {
        "全部类型": activity_current["totalAmount"],
        "经销商搭赠": activity_current["dealerAmount"],
        "消费者搭赠": activity_current["consumerGiftAmount"],
        "消费者试吃": activity_current["tasteAmount"],
    }
    sample_rank_columns = {
        "全部类型": sample_current["totalAmount"],
        "老客户增加品类送样": sample_current["oldAmount"],
        "新客户招商送样": sample_current["newAmount"],
    }

    return {
        "month": activity_current["month"] or sample_current["month"],
        "previousMonth": activity_previous["month"] or sample_previous["month"],
        "periodNote": "数据展示为T-1月",
        "source": "市场稽核部重点工作.xlsx / 赠品稽核-活动、赠品稽核-样品",
        "activity": {
            "count": clean_total(activity_count),
            "amount": clean_total(activity_amount),
            "countMom": mom_rate(activity_count, previous_activity_count),
            "amountMom": mom_rate(activity_amount, previous_activity_amount),
            "categories": {
                "经销商搭赠": category_payload(
                    activity_ws,
                    activity_current,
                    activity_previous,
                    "dealerCount",
                    "dealerAmount",
                ),
                "消费者搭赠": category_payload(
                    activity_ws,
                    activity_current,
                    activity_previous,
                    "consumerGiftCount",
                    "consumerGiftAmount",
                ),
                "消费者试吃": category_payload(
                    activity_ws,
                    activity_current,
                    activity_previous,
                    "tasteCount",
                    "tasteAmount",
                ),
            },
            "rankings": {
                label: ranked_amounts(
                    activity_ws, activity_current["name"], amount_column
                )
                for label, amount_column in activity_rank_columns.items()
            },
        },
        "sample": {
            "count": clean_total(sample_count),
            "amount": clean_total(sample_amount),
            "countMom": mom_rate(sample_count, previous_sample_count),
            "amountMom": mom_rate(sample_amount, previous_sample_amount),
            "categories": {
                "老客户增加品类送样": category_payload(
                    sample_ws,
                    sample_current,
                    sample_previous,
                    "oldCount",
                    "oldAmount",
                ),
                "新客户招商送样": category_payload(
                    sample_ws,
                    sample_current,
                    sample_previous,
                    "newCount",
                    "newAmount",
                ),
            },
            "customerNew": {
                "count": clean_total(customer_new_count),
                "countMom": None,
            },
            "rankings": {
                label: ranked_amounts(
                    sample_ws, sample_current["name"], amount_column
                )
                for label, amount_column in sample_rank_columns.items()
            },
        },
    }


def build_device_ban(wb):
    ws = wb[DEVICE_BAN_SHEET]
    grouped = defaultdict(
        lambda: {
            "total": 0,
            "statuses": Counter(),
            "provinces": Counter(),
            "types": Counter(),
            "platforms": Counter(),
        }
    )
    for row in range(3, ws.max_row + 1):
        month = month_label(ws.cell(row, 1).value) or month_label(ws.cell(row, 3).value)
        count = number(ws.cell(row, 8).value)
        if not month or count <= 0:
            continue
        bucket = grouped[month]
        bucket["total"] += count
        for key, column in (
            ("types", 2),
            ("platforms", 4),
            ("provinces", 5),
            ("statuses", 12),
        ):
            label = norm(ws.cell(row, column).value)
            if label:
                bucket[key][label] += count

    # P/Q and S/T contain the workbook's prepared status and province summaries.
    # Use them for the latest month when they reconcile to the underlying H-column total.
    if grouped:
        latest_month = max(grouped, key=month_key)
        latest = grouped[latest_month]
        status_summary = Counter()
        province_summary = Counter()
        for row in range(3, ws.max_row + 1):
            status = norm(ws.cell(row, 16).value)
            status_count = number(ws.cell(row, 17).value)
            if status and status != "总计" and status_count:
                status_summary[status] += status_count
            province = norm(ws.cell(row, 19).value)
            province_count = number(ws.cell(row, 20).value)
            if province and province_count:
                province_summary[province] += province_count
        if status_summary and sum(status_summary.values()) == latest["total"]:
            latest["statuses"] = status_summary
        if province_summary and sum(province_summary.values()) == latest["total"]:
            latest["provinces"] = province_summary

    payload = {}
    for month, bucket in grouped.items():
        status_rows = sorted(
            bucket["statuses"].items(), key=lambda item: (-item[1], item[0])
        )
        province_rows = sorted(
            bucket["provinces"].items(), key=lambda item: (-item[1], item[0])
        )
        type_rows = sorted(bucket["types"].items(), key=lambda item: (-item[1], item[0]))
        platform_rows = sorted(
            bucket["platforms"].items(), key=lambda item: (-item[1], item[0])
        )
        status_map = dict(status_rows)
        payload[month] = {
            "total": bucket["total"],
            "platform": "、".join(name for name, _ in platform_rows) or "--",
            "removed": status_map.get("已下架", 0),
            "noReply": status_map.get("未回复", 0),
            "reported": status_map.get("已向平台举报", 0),
            # 源表存在“获取商家信息”和“商家信息获取中”两种填写，均归入商家信息获取状态。
            "merchantInfo": status_map.get("获取商家信息", 0) + status_map.get("商家信息获取中", 0),
            "followUp": status_map.get("当地业务跟进中", 0),
            "topProvince": province_rows[0][0] if province_rows else "--",
            "topType": type_rows[0][0] if type_rows else "--",
            "statuses": [
                {"name": name, "count": count} for name, count in status_rows
            ],
            "provinces": [
                {"name": name, "count": count} for name, count in province_rows
            ],
            "types": [{"name": name, "count": count} for name, count in type_rows],
        }
    return payload


def build_store_audit(wb):
    ws = wb[STORE_AUDIT_SHEET]
    source_rows = list(
        ws.iter_rows(min_row=2, min_col=1, max_col=25, values_only=True)
    )
    years = [number(row[0]) for row in source_rows if number(row[0]) > 0]
    if not years:
        return {}
    latest_year = max(years)
    rows_by_month = defaultdict(list)
    for row in source_rows:
        if number(row[0]) != latest_year:
            continue
        month_number = number(row[1])
        if 1 <= month_number <= 12:
            rows_by_month[f"{month_number}月"].append(row)

    def store_text(value):
        text = norm(value)
        if not text:
            return ""
        try:
            return text.encode("latin1").decode("gbk")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return text

    def display_province(value):
        text = store_text(value)
        for suffix in (
            "壮族自治区",
            "回族自治区",
            "维吾尔自治区",
            "自治区",
            "特别行政区",
            "省",
            "市",
        ):
            if text.endswith(suffix):
                return text[: -len(suffix)]
        return text

    def top_label(counter):
        ranked = sorted(counter.items(), key=lambda item: (-item[1], item[0]))
        return ranked[0][0] if ranked else "--"

    payload = {}
    for month, rows in rows_by_month.items():
        results = Counter(store_text(row[21]) for row in rows)
        province_total = Counter(store_text(row[6]) for row in rows if store_text(row[6]))
        province_qualified = Counter(
            store_text(row[6])
            for row in rows
            if store_text(row[6]) and store_text(row[21]) == "合格"
        )
        province_uncertain = Counter(
            store_text(row[6])
            for row in rows
            if store_text(row[6]) and store_text(row[21]) == "无法判定"
        )
        province_bad = Counter(
            store_text(row[6])
            for row in rows
            if store_text(row[6]) and store_text(row[21]) == "不合格"
        )
        # 照片不规范按源表 V 列（判定结果）的“无法判定”口径统计，按省区汇总。
        province_photo = Counter(
            store_text(row[6])
            for row in rows
            if store_text(row[6]) and store_text(row[21]) == "无法判定"
        )
        province_no_product = Counter(
            store_text(row[6])
            for row in rows
            if store_text(row[6]) and store_text(row[22]) == "无小虎产品售卖"
        )
        sku_values = [
            float(row[10])
            for row in rows
            if isinstance(row[10], (int, float)) and not isinstance(row[10], bool)
        ]
        freezer_values = [
            float(row[11])
            for row in rows
            if isinstance(row[11], (int, float)) and not isinstance(row[11], bool)
        ]
        provinces = []
        for province, total in sorted(
            province_total.items(), key=lambda item: (-item[1], item[0])
        ):
            qualified = province_qualified[province]
            uncertain = province_uncertain[province]
            provinces.append(
                {
                    "n": display_province(province),
                    "t": total,
                    "q": qualified,
                    "u": uncertain,
                    "r": round(qualified / total * 100, 1) if total else 0,
                }
            )
        payload[month] = {
            "total": len(rows),
            "q": results["合格"],
            "bad": results["不合格"],
            "u": results["无法判定"],
            "avgSku": round(sum(sku_values) / len(sku_values), 1)
            if sku_values
            else 0,
            "avgFreezer": round(sum(freezer_values) / len(freezer_values), 1)
            if freezer_values
            else 0,
            "topBad": top_label(province_bad),
            "topPhoto": top_label(province_photo),
            "topUncertain": top_label(province_uncertain),
            "topNoProduct": top_label(province_no_product),
            "topNoProductCount": max(province_no_product.values(), default=0),
            "provinces": provinces,
        }
    return payload


def write_js(path: Path, legacy_name: str, map_name: str, payload_by_month: dict, default_month: str, extra=""):
    payload = payload_by_month.get(default_month) or next(iter(payload_by_month.values()), {})
    text = (
        f"window.{map_name} = "
        + json.dumps(payload_by_month, ensure_ascii=False, indent=2)
        + ";\n"
        + f"window.{legacy_name} = window.{map_name}[window.MAIN_SELECTED_MONTH || {json.dumps(default_month, ensure_ascii=False)}] || "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n"
        + extra
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def write_plain_js(path: Path, name: str, payload: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        f"window.{name} = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )


def bump_index_data_cache():
    index_path = ROOT / "index.html"
    if not index_path.exists():
        return
    html = index_path.read_text(encoding="utf-8")
    cache_tag = datetime.now().strftime("%Y%m%d%H%M%S")
    updated = re.sub(
        r'(<script\s+src="assets/data/[^"?]+\.js)(?:\?v=[^"]*)?("\s*></script>)',
        rf'\1?v={cache_tag}\2',
        html,
    )
    if updated != html:
        index_path.write_text(updated, encoding="utf-8")


def main():
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    promo_plan, promo_detail, promo_months = build_promo(wb)
    approval_pies, approval_detail, approval_months = build_approval(wb)
    market_order, market_months = build_market_order(wb)
    gift_audit = build_gift_audit(wb)
    device_ban = build_device_ban(wb)
    store_audit = build_store_audit(wb)
    source_months = sorted(
        set(promo_months)
        | set(approval_months)
        | set(market_months)
        | set(device_ban)
        | set(store_audit),
        key=month_key,
    )
    if not source_months:
        source_months = [month_label(datetime.now())]

    month_maps = (promo_plan, approval_pies, market_order, device_ban, store_audit)
    coverage = {
        month: sum(1 for payload_by_month in month_maps if month in payload_by_month)
        for month in source_months
    }
    default_month = max(source_months, key=lambda month: (coverage[month], month_key(month)))
    selector_months = [f"{month}月" for month in range(1, 13)]
    device_status, device_detail = build_device(wb, source_months)
    month_extra = (
        "window.MAIN_WORK_MONTHS = "
        + json.dumps(selector_months, ensure_ascii=False)
        + ";\nwindow.MAIN_DEFAULT_MONTH = "
        + json.dumps(default_month, ensure_ascii=False)
        + ";\n"
    )

    write_js(DATA_DIR / "promo-plan-audit.js", "PROMO_PLAN_AUDIT", "PROMO_PLAN_AUDIT_BY_MONTH", promo_plan, default_month, month_extra)
    write_js(DATA_DIR / "promo-audit-detail.js", "PROMO_AUDIT_DETAIL", "PROMO_AUDIT_DETAIL_BY_MONTH", promo_detail, default_month)
    write_js(DATA_DIR / "approval-pies.js", "APPROVAL_PIES", "APPROVAL_PIES_BY_MONTH", approval_pies, default_month)
    write_js(DATA_DIR / "approval-detail.js", "APPROVAL_DETAIL", "APPROVAL_DETAIL_BY_MONTH", approval_detail, default_month)
    write_js(DATA_DIR / "device-channel-status.js", "DEVICE_CHANNEL_STATUS", "DEVICE_CHANNEL_STATUS_BY_MONTH", device_status, default_month)
    write_js(DATA_DIR / "device-detail.js", "DEVICE_DETAIL", "DEVICE_DETAIL_BY_MONTH", device_detail, default_month)
    write_js(DATA_DIR / "market-order-governance.js", "MARKET_ORDER_GOVERNANCE", "MARKET_ORDER_GOVERNANCE_BY_MONTH", market_order, default_month)
    write_plain_js(DATA_DIR / "gift-audit.js", "GIFT_AUDIT", gift_audit)
    write_js(DATA_DIR / "device-ban-action.js", "DEVICE_BAN_ACTION", "DEVICE_BAN_ACTION_BY_MONTH", device_ban, default_month)
    write_plain_js(DATA_DIR / "store-audit-popup.js", "STORE_AUDIT_POPUP_BY_MONTH", store_audit)
    bump_index_data_cache()

    coverage_text = ", ".join(f"{month}:{coverage[month]}" for month in source_months)
    print(f"Updated month-aware work data: {', '.join(source_months)}; coverage {coverage_text}; default {default_month}")


if __name__ == "__main__":
    main()
