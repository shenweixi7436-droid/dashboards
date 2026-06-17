from collections import Counter
from datetime import datetime
from pathlib import Path
import json

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = Path("C:/Users/shenw/Desktop/\u770b\u677f/\u5e02\u573a\u7a3d\u6838\u90e8\u91cd\u70b9\u5de5\u4f5c.xlsx")
SHEET = "\u7ebf\u4e0a\u5ba1\u6279\u6d41\u7a0b\u7a3d\u6838\u660e\u7ec6"

OUT_DETAIL = ROOT / "assets" / "data" / "approval-detail.js"
OUT_PIES = ROOT / "assets" / "data" / "approval-pies.js"

DETAIL_COLS = [2, 4, 5, 6, 7, 8, 10]
RESULT_COL = 9
HEADER_ROW = 2
DATA_START_ROW = 3
FALLBACK_HEADERS = [
    "\u6d41\u7a0b\u53d1\u8d77\u65f6\u95f4",
    "\u8d39\u7528\u7c7b\u578b",
    "\u7b7e\u5448\u53f7",
    "\u7701\u533a",
    "\u5ba2\u6237\u540d\u79f0",
    "\u53d1\u8d77\u4eba",
    "\u95ee\u9898\u7c7b\u578b",
]


def norm(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def main():
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb[SHEET]

    headers = [
        norm(ws.cell(HEADER_ROW, col).value) or fallback
        for fallback, col in zip(FALLBACK_HEADERS, DETAIL_COLS)
    ]

    total = 0
    qualified = 0
    unqualified = 0
    rows = []
    issue_counter = Counter()

    needed_cols = sorted(set(DETAIL_COLS + [RESULT_COL]))
    max_needed_col = max(needed_cols)

    for row_idx, raw_row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, max_col=max_needed_col, values_only=True),
        DATA_START_ROW,
    ):
        detail_values = [
            norm(raw_row[col - 1] if len(raw_row) >= col else "")
            for col in DETAIL_COLS
        ]
        result = norm(raw_row[RESULT_COL - 1] if len(raw_row) >= RESULT_COL else "")

        if not any(detail_values) and not result:
            continue

        total += 1
        if result == "\u662f":
            qualified += 1
        else:
            unqualified += 1
            issue_name = detail_values[-1]
            if issue_name:
                issue_counter[issue_name] += 1
            rows.append(
                {
                    "row": row_idx,
                    "values": detail_values,
                    "result": "\u4e0d\u5408\u683c",
                }
            )

    detail_payload = {
        "month": 6,
        "source": "\u5e02\u573a\u7a3d\u6838\u90e8\u91cd\u70b9\u5de5\u4f5c.xlsx / \u7ebf\u4e0a\u5ba1\u6279\u6d41\u7a0b\u7a3d\u6838\u660e\u7ec6",
        "headers": headers,
        "total": total,
        "qualified": qualified,
        "unqualified": unqualified,
        "rows": rows,
    }
    pie_payload = {
        "total": total,
        "qualified": qualified,
        "unqualified": unqualified,
        "rate": round(qualified / total * 100, 1) if total else 0,
        "issues": [
            {"name": name, "value": value}
            for name, value in issue_counter.most_common(3)
        ],
    }

    OUT_DETAIL.parent.mkdir(parents=True, exist_ok=True)
    OUT_DETAIL.write_text(
        "window.APPROVAL_DETAIL = "
        + json.dumps(detail_payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    OUT_PIES.write_text(
        "window.APPROVAL_PIES = "
        + json.dumps(pie_payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    print(f"Generated: {OUT_DETAIL}")
    print(f"Generated: {OUT_PIES}")
    print(f"Approval total: {total}, qualified: {qualified}, unqualified: {unqualified}")


if __name__ == "__main__":
    main()
