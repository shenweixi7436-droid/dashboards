# -*- coding: utf-8 -*-
"""重建 store-business-analysis.html 内嵌的 RAW_DATA 明细快照。

用途：替代已丢失的 build_audit_dashboard.py 中更新鸣忙专项稽核看板页面的部分。
数据源：市场稽核部重点工作.xlsx 的"鸣忙门店专项稽核"sheet。
用法：
    python build_store_audit_dashboard.py            # 更新 RAW_DATA
    python build_store_audit_dashboard.py --check    # 仅校验：与现有快照逐行比对，不写文件
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SOURCE = Path("C:/Users/shenw/Desktop/看板/市场稽核部重点工作.xlsx")
PAGE = ROOT / "store-business-analysis.html"
SHEET = "鸣忙门店专项稽核"
PHOTO_URL_PREFIX = "门店经营分析_照片"

# 列号(1-based)：与 RAW_DATA 字段的映射，依据原快照逐行比对确认
COL_YEAR = 1        # 年清洗
COL_MONTH = 2       # 月份清洗
COL_STORE = 6       # 门店名称
COL_PROVINCE = 7    # 省份
COL_CITY = 8        # 城市（同时作为 auditGroup）
COL_DISTRICT = 9    # 区县
COL_ADDRESS = 10    # 门店地址
COL_SKU = 11        # 陈列SKU数
COL_FREEZER = 12    # 陈列冰柜数
COL_SAUSAGE = 13    # 烤肠零售价
COL_SHELL = 14      # 蛋挞皮零售价
COL_LIQUID = 15     # 蛋挞液零售价
COL_COMPETITOR = 16 # 竞品记录信息
COL_IMPROVE = 17    # 本次巡店改进项
COL_REGION = 18     # 省区
COL_DATE = 19       # 巡店日期
COL_REG_AT = 20     # 登记时间
COL_REGISTRAR = 21  # 登记人
COL_RESULT = 22     # 判定结果
COL_REASON = 23     # 不合格原因
COL_PHOTO_OK = 24   # 冰柜照片是否完整
COL_REMARK = 25     # 备注

DISPIMG_RE = re.compile(r'DISPIMG\("([^"]+)"')


def text_value(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    # 源表中部分单元格为 UTF-8 字节被按 GBK 解码产生的乱码（如"闄堝垪涓嶈冻鍗婃煖"），
    # 反向修复：先按 GBK 编回字节，再按 UTF-8 解码；失败则保留原文。
    try:
        repaired = text.encode("gbk").decode("utf-8")
        if repaired and "\ufffd" not in repaired:
            return repaired
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    return text


def date_value(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return str(value)


def main():
    check_only = "--check" in sys.argv
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb[SHEET]

    # 逐行构建（保留 Excel 真实行号作为 rowNum）
    out = []
    for excel_row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_col=25, values_only=True), start=2
    ):
        if row[COL_YEAR - 1] in (None, ""):
            continue
        remark = text_value(row[COL_REMARK - 1])
        photo_ids = DISPIMG_RE.findall(remark)
        result = text_value(row[COL_RESULT - 1]) or "未填写"
        reason = text_value(row[COL_REASON - 1])
        if result == "合格" or reason in ("/", ""):
            # 原生成器口径：合格一律记"无不合格原因"；未填写原因记"未填写原因"
            reason = "无不合格原因" if result == "合格" else "未填写原因"

        def num(v):
            return v if v not in (None, "") else ""

        out.append(
            {
                "rowNum": excel_row_idx,
                "yearClean": text_value(row[COL_YEAR - 1]),
                "monthClean": text_value(row[COL_MONTH - 1]),
                "auditGroup": text_value(row[COL_CITY - 1]),
                "store": text_value(row[COL_STORE - 1]),
                "province": text_value(row[COL_PROVINCE - 1]),
                "city": text_value(row[COL_CITY - 1]),
                "district": text_value(row[COL_DISTRICT - 1]),
                "address": text_value(row[COL_ADDRESS - 1]),
                "sku": num(row[COL_SKU - 1]),
                "freezer": num(row[COL_FREEZER - 1]),
                "sausagePrice": num(row[COL_SAUSAGE - 1]),
                "eggTartShellPrice": num(row[COL_SHELL - 1]),
                "eggTartLiquidPrice": num(row[COL_LIQUID - 1]),
                "competitor": text_value(row[COL_COMPETITOR - 1]),
                "improvement": text_value(row[COL_IMPROVE - 1]),
                "region": text_value(row[COL_REGION - 1]),
                "date": date_value(row[COL_DATE - 1]),
                "registeredAt": date_value(row[COL_REG_AT - 1]),
                "registrar": text_value(row[COL_REGISTRAR - 1]),
                "result": result,
                "reason": reason,
                "photoComplete": text_value(row[COL_PHOTO_OK - 1]),
                "remark": remark,
                "photoIds": photo_ids,
                "photoUrls": [f"{PHOTO_URL_PREFIX}/{pid}.jpeg" for pid in photo_ids],
                "hasPhoto": bool(photo_ids),
            }
        )

    print(f"built rows from Excel: {len(out)}")

    html = PAGE.read_text(encoding="utf-8")
    m = re.search(r"var RAW_DATA = (\[.*?\]);\s*\n", html, re.S)
    if not m:
        print("ERROR: RAW_DATA block not found in page")
        sys.exit(1)
    existing = json.loads(m.group(1))
    print(f"existing snapshot rows: {len(existing)}")

    if check_only:
        old_by_num = {r["rowNum"]: r for r in existing}
        mismatch = 0
        for i, new in enumerate(out):
            old = old_by_num.get(new["rowNum"])
            if old is None:
                continue
            if old != new:
                mismatch += 1
                if mismatch <= 3:
                    for k in new:
                        if old.get(k) != new[k]:
                            print(
                                f"rowNum={new['rowNum']} field={k}: "
                                f"snapshot={old.get(k)!r} excel={new[k]!r}"
                            )
        common = sum(1 for r in out if r["rowNum"] in old_by_num)
        print(f"compared {common} common rows, mismatches: {mismatch}")
        sys.exit(0 if mismatch == 0 else 2)

    new_json = json.dumps(out, ensure_ascii=False, separators=(",", ":"))
    block = f"var RAW_DATA = {new_json};\n"
    html = html[: m.start()] + block + html[m.end():]

    # 同步更新页头"内置数据：N 条"计数
    pill_new = f"内置数据：{len(out):,} 条"
    html, n_pill = re.subn(
        r"内置数据：[\d,]+ 条", pill_new, html, count=1
    )

    PAGE.write_text(html, encoding="utf-8")
    print(f"RAW_DATA updated in {PAGE.name}: {len(out)} rows (source-pill updated: {n_pill})")


if __name__ == "__main__":
    main()
