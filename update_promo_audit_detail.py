from collections import Counter
from pathlib import Path
import json

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = Path("C:/Users/shenw/Desktop/\u770b\u677f/\u5e02\u573a\u7a3d\u6838\u90e8\u91cd\u70b9\u5de5\u4f5c.xlsx")

AUDIT_SHEET = "\u63a8\u5e7f\u4fc3\u9500\u7a3d\u6838"
PLAN_SHEET = "\u63a8\u5e7f\u4fc3\u9500\u8ba1\u5212"

OUT_DETAIL = ROOT / "assets" / "data" / "promo-audit-detail.js"
OUT_PLAN = ROOT / "assets" / "data" / "promo-plan-audit.js"

DETAIL_COLS = [2, 3, 4, 6, 14]
RESULT_COL = 14
HEADER_ROW = 3
DATA_START_ROW = 4
FALLBACK_HEADERS = [
    "\u7701\u4efd",
    "\u5e02\u533a",
    "\u57ce\u5e02\u7ecf\u7406",
    "\u6d3b\u52a8\u95e8\u5e97\u540d\u79f0",
    "\u7a3d\u6838\u7ed3\u679c",
]

PLAN_PROVINCE_COL = 3
PLAN_TARGET_COL = 5
PLAN_START_ROW = 2


def norm(value):
    if value is None:
        return ""
    return str(value).strip()


def number(value):
    if value is None or value == "":
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def result_label(value):
    text = norm(value)
    return "\u5408\u683c" if text in {"\u662f", "\u5408\u683c"} else "\u4e0d\u5408\u683c"


def main():
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws_audit = wb[AUDIT_SHEET]
    ws_plan = wb[PLAN_SHEET]

    headers = [
        norm(ws_audit.cell(HEADER_ROW, col).value) or fallback
        for fallback, col in zip(FALLBACK_HEADERS, DETAIL_COLS)
    ]

    rows = []
    audit_by_province = Counter()
    max_col = max(DETAIL_COLS + [RESULT_COL])

    for row_idx, raw_row in enumerate(
        ws_audit.iter_rows(min_row=DATA_START_ROW, max_col=max_col, values_only=True),
        DATA_START_ROW,
    ):
        values = [norm(raw_row[col - 1] if len(raw_row) >= col else "") for col in DETAIL_COLS]
        result = norm(raw_row[RESULT_COL - 1] if len(raw_row) >= RESULT_COL else "")
        if not any(values) and not result:
            continue

        province = values[0]
        if result:
            audit_by_province[province] += 1

        rows.append(
            {
                "row": row_idx,
                "values": values,
                "result": result_label(result),
            }
        )

    total = len(rows)
    qualified = sum(1 for row in rows if row["result"] == "\u5408\u683c")
    unqualified = total - qualified

    plan_by_province = {}
    for row_idx in range(PLAN_START_ROW, ws_plan.max_row + 1):
        province = norm(ws_plan.cell(row_idx, PLAN_PROVINCE_COL).value)
        if not province:
            continue
        plan_by_province[province] = plan_by_province.get(province, 0) + number(
            ws_plan.cell(row_idx, PLAN_TARGET_COL).value
        )

    province_order = list(plan_by_province.keys())
    for province in audit_by_province.keys():
        if province and province not in plan_by_province:
            province_order.append(province)

    plan_total = sum(plan_by_province.values())
    audit_total = sum(audit_by_province.values())
    progress = round(audit_total / plan_total * 100, 1) if plan_total else 0
    plan_rows = [
        {
            "province": province,
            "plan": int(plan_by_province.get(province, 0)),
            "audit": int(audit_by_province.get(province, 0)),
        }
        for province in province_order
    ]

    detail_payload = {
        "month": 6,
        "source": "\u5e02\u573a\u7a3d\u6838\u90e8\u91cd\u70b9\u5de5\u4f5c.xlsx / \u63a8\u5e7f\u4fc3\u9500\u7a3d\u6838",
        "headers": headers,
        "total": total,
        "qualified": qualified,
        "unqualified": unqualified,
        "rows": rows,
    }
    plan_payload = {
        "planTotal": plan_total,
        "auditTotal": audit_total,
        "progress": progress,
        "rows": plan_rows,
    }

    OUT_DETAIL.parent.mkdir(parents=True, exist_ok=True)
    OUT_DETAIL.write_text(
        "window.PROMO_AUDIT_DETAIL = "
        + json.dumps(detail_payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    OUT_PLAN.write_text(
        "window.PROMO_PLAN_AUDIT = "
        + json.dumps(plan_payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    print(f"Generated: {OUT_DETAIL}")
    print(f"Generated: {OUT_PLAN}")
    print(f"Promo audit total: {total}, qualified: {qualified}, unqualified: {unqualified}")
    print(f"Promo task total: {plan_total}, audited: {audit_total}, progress: {progress}%")


if __name__ == "__main__":
    main()
